from django.shortcuts import render,get_object_or_404,redirect
from control.models import Pedidos_Venda
from usuarios.models import Usuario
from . forms import PedidosVendaForm
from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from control.utils import registrar_acao, ultimas_acoes_modulo
from django.contrib import messages

#PEDIDOS
@login_required
def pedidos(request):

    # Total de pedidos:
    total_pedidos = Pedidos_Venda.objects.filter(status__in=['aberto', 'processando']).count()

    search = request.GET.get("search")

    pedidos = Pedidos_Venda.objects.filter(status__in=['aberto', 'processando'])

    if search:
        pedidos = pedidos.filter(
            Q(produto__nome__icontains=search) |
            Q(cliente__nome__icontains=search) |
            Q(id__icontains=search) 
        )
    
    pedidos = pedidos.order_by('-data')

    paginator = Paginator(pedidos, 10)
    numero_da_pagina = request.GET.get('p')
    pedidos_paginados = paginator.get_page(numero_da_pagina)
    return render(request, "vendas/pedidos/pedidos.html", {"pedidos": pedidos_paginados, "search": search, "total_pedidos": total_pedidos, 'historico': ultimas_acoes_modulo(request.user, 'pedidos_venda')})

@login_required
def registrar_pedido(request):
    if request.method == "POST":
        form = PedidosVendaForm(request.POST)
        if form.is_valid():
            pedidos_venda = form.save()
            messages.success(
                request,
                f"Pedido de {pedidos_venda.produto.nome} registrado com sucesso."
            )

            registrar_acao(
                request.user,
                'pedidos_venda',
                f"Pedido de {pedidos_venda.produto.nome} <strong>registrado</strong>"
            )
            return redirect("vendas:pedidos")
        else:
            print("Erros de Validação:", form.errors)
    else:
        form = PedidosVendaForm()
    return render(request, "vendas/pedidos/registrar_pedido.html", {"form": form, 'historico': ultimas_acoes_modulo(request.user, 'pedidos_venda')})

@login_required
def editar_pedido(request, id_pedido):
    pedido = get_object_or_404(Pedidos_Venda, id=id_pedido)

    if request.method == "POST":
        form = PedidosVendaForm(request.POST, instance=pedido)
        if form.is_valid():
            pedidos_venda = form.save()
            messages.success(
                request,
                f"Pedido de {pedidos_venda.produto.nome} editado com sucesso."
            )

            registrar_acao(
                request.user,
                'pedidos_venda',
                f"{pedidos_venda.produto.nome} <strong>editado</strong>"
            )
            return redirect("vendas:pedidos")
    else:
        form = PedidosVendaForm(instance=pedido)

    return render(request, "vendas/pedidos/editar_pedido.html", {"form": form, 'historico': ultimas_acoes_modulo(request.user, 'pedidos_venda'), "pedido": pedido})
    
@login_required
def excluir_pedido(request, id_pedido=0):
    if request.method == "POST":
        ids = request.POST.getlist("ids_selecionados")

        quantidade = len(ids)

        if quantidade > 0:
            Pedidos_Venda.objects.filter(id__in=ids).delete() 
            messages.success(
                request,
                f"Pedido excluído com sucesso."
            )
            # REGISTRA A AÇÃO
            registrar_acao(
                usuario=request.user,
                modulo='pedidos_venda',
                descricao=f'{quantidade} pedido(s) excluído(s)'
            )
        return redirect('vendas:pedidos')
    else:
        pedido = get_object_or_404(Pedidos_Venda, id=id_pedido)
        return render(request, "vendas/pedidos/confirma.html", {"pedido": pedido, 'historico': ultimas_acoes_modulo(request.user, 'pedidos_venda')})
    
#HISTÓRICO
@login_required
def historico(request):

    # Vendas concluídas:
    total_vendas = Pedidos_Venda.objects.filter(status='concluido').count()

    search = request.GET.get("search")

    pedidos = Pedidos_Venda.objects.filter(status__in=['concluido', 'cancelado'])

    if search:
        pedidos = pedidos.filter(
            Q(produto__nome__icontains=search) |
            Q(cliente__nome__icontains=search) 
        )

    pedidos = pedidos.order_by('-data')

    paginator = Paginator(pedidos, 10)
    numero_da_pagina = request.GET.get('p')
    pedidos_paginados = paginator.get_page(numero_da_pagina)
    return render(request, "vendas/historico/historico.html", {"pedidos": pedidos_paginados, "search": search, "total_vendas": total_vendas})

#RELATÓRIO DE VENDAS
@login_required
def relatorio_vendas(request):

    # Vendas concluídas:
    total_vendas = Pedidos_Venda.objects.filter(status='concluido').count()

    ano = request.GET.get("ano")
    mes = request.GET.get("mes")
    search = request.GET.get("search")

    pedidos = Pedidos_Venda.objects.filter(status__in=['concluido'])

    if ano and ano.isdigit():
        pedidos = pedidos.filter(data__year=ano)

    if mes and mes.isdigit():
        pedidos = pedidos.filter(data__month=mes)


    if search:
        pedidos = pedidos.filter(
            Q(produto__nome__icontains=search) |
            Q(cliente__nome__icontains=search) |
            Q(vendedor__nome__icontains=search) 
        )

    pedidos = pedidos.order_by('-data')

    anos_disponiveis = Pedidos_Venda.objects.dates('data', 'year')

    anos = [a.year for a in anos_disponiveis]
    anos.sort(reverse=True)
    
    meses_lista = [
        ("1", "Janeiro"), ("2", "Fevereiro"), ("3", "Março"),
        ("4", "Abril"), ("5", "Maio"), ("6", "Junho"),
        ("7", "Julho"), ("8", "Agosto"), ("9", "Setembro"),
        ("10", "Outubro"), ("11", "Novembro"), ("12", "Dezembro")
    ]

    paginator = Paginator(pedidos, 10)
    numero_da_pagina = request.GET.get('p')
    pedidos_paginados = paginator.get_page(numero_da_pagina)
    total_vendas = Pedidos_Venda.objects.filter(status='concluido').count()
    return render(request, "vendas/relatorio/relatorio_vendas.html", {"pedidos": pedidos_paginados, "anos": anos, "meses": meses_lista, "ano_selecionado": ano, "mes_selecionado": mes, "search": search, "total_vendas": total_vendas})

@login_required
def exportar_pdf(request):
    ano = request.GET.get("ano")
    mes = request.GET.get("mes")
    search = request.GET.get("search")

    pedidos = Pedidos_Venda.objects.all()

    #FILTROS
    if ano:
        pedidos = pedidos.filter(data__year=ano)
    if mes:
        pedidos = pedidos.filter(data__month=mes)
    if search:
        pedidos = pedidos.filter(
            Q(produto__nome__icontains=search) |
            Q(cliente__nome__icontains=search) |
            Q(vendedor__nome__icontains=search)
        )

    pedidos = pedidos.order_by("-data")
    
    nome_mes = None
    if mes:
        meses_pt = [
            "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
            "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
        ]
        mes_int = int(mes)
        nome_mes = meses_pt[mes_int - 1]

    #NOME DO ARQUIVO
    hoje = timezone.localtime().strftime("%d-%m-%Y")
    nome_arquivo = f"relatorio_vendas_{hoje}.pdf"

    #RESPOSTA HTTP
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'

    #CRIAR PDF
    pdf = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    #CABEÇALHO PROFISSIONAL
    cabecalho_style = ParagraphStyle(
        name="Cabecalho",
        fontSize=12,
        alignment=1,
        spaceAfter=6,
        leading=16,
    )

    razao_social = request.user.razao_social
    cnpj = request.user.cnpj
    email = request.user.email
    telefone = request.user.telefone

    elements.append(Paragraph(f"<b>{razao_social}</b>", cabecalho_style))
    elements.append(Paragraph(cnpj, cabecalho_style))
    elements.append(Paragraph(email, cabecalho_style))
    elements.append(Paragraph(telefone, cabecalho_style))
    elements.append(Spacer(1, 20))

    #TÍTULO
    titulo = Paragraph("<b>Relatório de Vendas</b>", styles["Title"])
    elements.append(titulo)

    #SUBTÍTULO DOS FILTROS
    sub_style = ParagraphStyle(
        name="Subtitulo",
        fontSize=14,
        alignment=1,
        spaceAfter=20,
    )

    #Texto dinâmico conforme filtro
    descricao_filtro = "Relatório Geral"

    if ano and mes:
        descricao_filtro = f"{nome_mes}/{ano}"
    elif ano:
        descricao_filtro = f"{ano}"
    elif mes:
        descricao_filtro = f"{nome_mes} (todos os anos)"

    elements.append(Paragraph(f"<b>{descricao_filtro}</b>", sub_style))

    #CABEÇALHO DA TABELA
    dados = [
        ["Produto", "Cliente", "Vendedor", "Preço", "Qtd", "Total", "Data"]
    ]

    #LINHAS DA TABELA
    for p in pedidos:
        dados.append([
            str(p.produto),
            str(p.cliente),
            str(p.vendedor),
            f"R$ {p.produto.preco:.2f}",
            p.quantidade,
            f"R$ {(p.produto.preco * p.quantidade):.2f}",
            timezone.localtime(p.data).strftime("%d/%m/%Y %H:%M"),
        ])

    #TABELA
    tabela = Table(dados)

    #ESTILOS DA TABELA
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),

        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),

        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),

        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    elements.append(tabela)

    #CONSTRUIR PDF
    pdf.build(elements)

    return response

@login_required
def visualizar_pdf(request):
    ano = request.GET.get("ano")
    mes = request.GET.get("mes")
    search = request.GET.get("search")

    pedidos = Pedidos_Venda.objects.all()

    if ano:
        pedidos = pedidos.filter(data__year=ano)
    if mes:
        pedidos = pedidos.filter(data__month=mes)
    if search:
        pedidos = pedidos.filter(
            Q(produto__nome__icontains=search) |
            Q(cliente__nome__icontains=search) |
            Q(vendedor__nome__icontains=search)
        )

    pedidos = pedidos.order_by("-data")

    nome_mes = None
    if mes:
        meses_pt = [
            "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
            "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
        ]
        mes_int = int(mes)
        nome_mes = meses_pt[mes_int - 1]

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "inline; filename=relatorio_vendas.pdf"

    pdf = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    #CABEÇALHO PROFISSIONAL
    cabecalho_style = ParagraphStyle(
        name="Cabecalho",
        fontSize=12,
        alignment=1,
        spaceAfter=6,
        leading=16,
    )

    razao_social = request.user.razao_social
    cnpj = request.user.cnpj
    email = request.user.email
    telefone = request.user.telefone

    elements.append(Paragraph(f"<b>{razao_social}</b>", cabecalho_style))
    elements.append(Paragraph(cnpj, cabecalho_style))
    elements.append(Paragraph(email, cabecalho_style))
    elements.append(Paragraph(telefone, cabecalho_style))
    elements.append(Spacer(1, 20))

    #TÍTULO
    titulo = Paragraph("<b>Relatório de Vendas</b>", styles["Title"])
    elements.append(titulo)

    #SUBTÍTULO DOS FILTROS
    sub_style = ParagraphStyle(
        name="Subtitulo",
        fontSize=14,
        alignment=1,
        spaceAfter=20,
    )

    #Texto dinâmico conforme filtro
    descricao_filtro = "Relatório Geral"

    if ano and mes:
        descricao_filtro = f"{nome_mes}/{ano}"
    elif ano:
        descricao_filtro = f"{ano}"
    elif mes:
        descricao_filtro = f"{nome_mes} (todos os anos)"

    elements.append(Paragraph(f"<b>{descricao_filtro}</b>", sub_style))

    #CABEÇALHO DA TABELA
    dados = [
        ["Produto", "Cliente", "Vendedor", "Preço", "Qtd", "Total", "Data"]
    ]

    #LINHAS DA TABELA
    for p in pedidos:
        dados.append([
            str(p.produto),
            str(p.cliente),
            str(p.vendedor),
            f"R$ {p.produto.preco:.2f}",
            p.quantidade,
            f"R$ {(p.produto.preco * p.quantidade):.2f}",
            timezone.localtime(p.data).strftime("%d/%m/%Y %H:%M"),
        ])

    #TABELA
    tabela = Table(dados)

    #ESTILOS DA TABELA
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),

        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),

        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),

        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    elements.append(tabela)

    pdf.build(elements)

    return response

@login_required
def exportar_excel(request):
    ano = request.GET.get("ano")
    mes = request.GET.get("mes")
    search = request.GET.get("search")

    pedidos = Pedidos_Venda.objects.all()

    #FILTROS
    if ano:
        pedidos = pedidos.filter(data__year=ano)
    if mes:
        pedidos = pedidos.filter(data__month=mes)
    if search:
        pedidos = pedidos.filter(
            Q(produto__nome__icontains=search) |
            Q(cliente__nome__icontains=search) |
            Q(vendedor__nome__icontains=search)
        )

    pedidos = pedidos.order_by('-data')

    #ARQUIVO
    hoje = timezone.now().strftime("%d-%m-%Y")
    nome_arquivo = f"relatorio_vendas_{hoje}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    #CABEÇALHO
    headers = ["Produto", "Cliente", "Vendedor", "Preço", "Quantidade", "Total", "Data"]

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(headers)

    #Formatando cabeçalho
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    #DADOS
    for pedido in pedidos:
        data_local = timezone.localtime(pedido.data)

        ws.append([
            str(pedido.produto),
            str(pedido.cliente),
            str(pedido.vendedor),
            float(pedido.produto.preco),
            pedido.quantidade,
            float(pedido.produto.preco * pedido.quantidade),
            data_local.strftime("%d/%m/%Y %H:%M"),
        ])

    #Ajuste de largura automática
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter

        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        ws.column_dimensions[column].width = max_length + 2

    #RESPOSTA
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'

    wb.save(response)
    return response
